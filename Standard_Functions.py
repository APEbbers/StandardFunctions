# *************************************************************************************
# *   MIT License                                                                     *
# *                                                                                   *
# *   Copyright (c) 2023 Paul Ebbers                                                  *
# *                                                                                   *
# *   Permission is hereby granted, free of charge, to any person obtaining a copy    *
# *   of this software and associated documentation files (the "Software"), to deal   *
# *   in the Software without restriction, including without limitation the rights    *
# *   to use, copy, modify, merge, publish, distribute, sublicense, and/or sell       *
# *   copies of the Software, and to permit persons to whom the Software is           *
# *   furnished to do so, subject to the following conditions:                        *
# *                                                                                   *
# *   The above copyright notice and this permission notice shall be included in all  *
# *   copies or substantial portions of the Software.                                 *
# *                                                                                   *
# *   THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR      *
# *   IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,        *
# *   FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE     *
# *   AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER          *
# *   LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,   *
# *   OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE   *
# *   SOFTWARE.                                                                       *
# *                                                                                   *
# *************************************************************************************/


class StandardFunctions_FreeCAD:
    @classmethod
    def Mbox(
        self,
        text,
        title="",
        style=0,
        IconType="Information",
        default="",
        stringList="[,]",
    ):
        """
        Message Styles:\n
        0 : OK                          (text, title, style)\n
        1 : Yes | No                    (text, title, style)\n
        20 : Inputbox                    (text, title, style, default)\n
        21 : Inputbox with dropdown      (text, title, style, default, stringlist)\n
        """
        from PySide.QtWidgets import QMessageBox, QInputDialog

        Icon = QMessageBox.Information
        if IconType == "NoIcon":
            Icon = QMessageBox.NoIcon
        if IconType == "Question":
            Icon = QMessageBox.Question
        if IconType == "Warning":
            Icon = QMessageBox.Warning
        if IconType == "Critical":
            Icon = QMessageBox.Critical

        if style == 0:
            # Set the messagebox
            msgBox = QMessageBox()
            msgBox.setIcon(Icon)
            msgBox.setText(text)
            msgBox.setWindowTitle(title)

            reply = msgBox.exec_()
            return reply
        if style == 1:
            # Set the messagebox
            msgBox = QMessageBox()
            msgBox.setIcon(Icon)
            msgBox.setText(text)
            msgBox.setWindowTitle(title)
            # Set the buttons and default button
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgBox.setDefaultButton(QMessageBox.No)

            reply = msgBox.exec_()
            if reply == QMessageBox.Yes:
                return "yes"
            if reply == QMessageBox.No:
                return "no"
        if style == 20:
            reply = QInputDialog.getText(
                parent=None, title=title, label=text, text=default
            )
            if reply[1]:
                # user clicked OK
                replyText = reply[0]
            else:
                # user clicked Cancel
                replyText = reply[0]  # which will be "" if they clicked Cancel
            return str(replyText)
        if style == 21:
            reply = QInputDialog.getItem(
                parent=None,
                title=title,
                label=text,
                items=stringList,
                current=1,
                editable=True,
            )
            if reply[1]:
                # user clicked OK
                replyText = reply[0]
            else:
                # user clicked Cancel
                replyText = reply[0]  # which will be "" if they clicked Cancel
            return str(replyText)

    @classmethod
    def GetFolder(self, parent=None, DefaultPath="") -> str:
        from PySide.QtWidgets import QFileDialog
    
        Directory = ""
        Directory = QFileDialog.getExistingDirectory(parent=parent, caption="Select Folder", dir=DefaultPath)

    return Directory
    
    @classmethod
    def GetFileDialog(self, Filter="", parent=None, DefaultPath="", SaveAs: bool = True) -> str:
    """
    Set filter like:
    "Images (*.png *.xpm .jpg);;Text files (.txt);;XML files (*.xml)"
    SaveAs:\n
        If True,  as SaveAs dialog will open and the file will be overwritten\n
        If False, an OpenFile dialog will be open and the file will be opened.\n
    """
    from PySide.QtWidgets import QFileDialog

    file = ""
    if SaveAs is False:
        file = QFileDialog.getOpenFileName(parent=parent, caption="Select a file", dir=DefaultPath, filter=Filter)[0]
    if SaveAs is True:
        file = QFileDialog.getSaveFileName(parent=parent, caption="Select a file", dir=DefaultPath, filter=Filter)[0]
    return file

    @classmethod
    def GetLetterFromNumber(self, number: int, UCase: bool = True):
        from openpyxl.utils import get_column_letter

        Letter = get_column_letter(number)

        # If UCase is true, convert to upper case
        if UCase is True:
            Letter = Letter.upper()

        return Letter

    @classmethod
    def GetNumberFromLetter(self, Letter):
        from openpyxl.utils import column_index_from_string

        Number = column_index_from_string(Letter)

        return Number

    @classmethod
    def GetA1fromR1C1(self, input: str) -> str:
        if input[:1] == "'":
            input = input[1:]
        try:
            input = input.upper()
            ColumnPosition = input.find("C")
            RowNumber = int(input[1:(ColumnPosition)])
            ColumnNumber = int(input[(ColumnPosition + 1) :])

            ColumnLetter = self.GetLetterFromNumber(ColumnNumber)

            return str(ColumnLetter + str(RowNumber))
        except Exception:
            return ""

    @classmethod
    def CheckIfWorkbookExists(self, FullFileName: str, CreateIfNone: bool = True):
        import os
        from openpyxl import Workbook

        result = False
        try:
            result = os.path.exists(FullFileName)
        except Exception:
            if CreateIfNone is True:
                Filter = [
                    ("Excel", "*.xlsx"),
                    (
                        "Excel Macro-enabled Workbook",
                        "*.xlsm",
                    ),
                ]
                FullFileName = SaveDialog(Filter)
                if FullFileName.strip():
                    wb = Workbook(str(FullFileName))
                    wb.save(FullFileName)
                    wb.close()
                    result = True
            if CreateIfNone is False:
                result = False
        return result

    @classmethod
    def ColorConvertor(self, ColorRGB: [], Alpha: float = 255) -> ():
        """
        A single function to convert colors to rgba colors as a tuple of float from 0-1
        ColorRGB:   [255,255,255]
        Alpha:      0-1
        """
        from matplotlib import colors as mcolors

        ColorRed = ColorRGB[0] / 255
        colorGreen = ColorRGB[1] / 255
        colorBlue = ColorRGB[2] / 255
        ColorAlpha = Alpha / 255

        color = (ColorRed, colorGreen, colorBlue)
        result = mcolors.to_rgba(color, ColorAlpha)

        return result

    @classmethod
    def OpenFile(self, FileName: str):
        """
        Filename = full path with filename as string
        """
        import subprocess
        import os
        import platform

        try:
            if os.path.exists(FileName):
                if platform.system() == "Darwin":  # macOS
                    subprocess.call(("open", FileName))
                elif platform.system() == "Windows":  # Windows
                    os.startfile(FileName)
                else:  # linux variants
                    print(FileName)
                    try:
                        subprocess.check_output(["xdg-open", FileName.strip()])
                    except subprocess.CalledProcessError:
                        Print(
                            f"An error occured when opening {FileName}!\n"
                            + "This can happen when running FreeCAD as an AppImage.\n"
                            + "Please install FreeCAD directly.",
                            "Error",
                        )
            else:
                print(f"Error: {FileName} does not exist.")
        except Exception as e:
            raise e

    @classmethod
    def SetColumnWidth_SpreadSheet(
        self, sheet, column: str, cellValue: str, factor: int = 10
    ) -> bool:
        """_summary_

        Args:
            sheet (_type_): FreeCAD spreadsheet object.\n
            column (str): The column for which the width will be set. must be like "A", "B", etc.\n
            cellValue (str): The string to calulate the widht from.\n
            factor (int, optional): to increase the stringlength with a factor. Defaults to 10.\n

        Returns:
            bool: returns True or False
        """
        try:
            # Calculate the text length needed.
            length = int(len(cellValue) * factor)

            print(column)
            # Set the column width
            sheet.setColumnWidth(column, length)

            # Recompute the sheet
            sheet.recompute()
        except Exception:
            return False

        return True

    @classmethod
    def Print(self, Input: str, Type: str = ""):
        """_summary_

        Args:
            Input (str): Text to print.\n
            Type (str, optional): Type of message. (enter Warning, Error or Log). Defaults to "".
        """
        import FreeCAD as App

        if Type == "Warning":
            App.Console.PrintWarning(Input + "\n")
        elif Type == "Error":
            App.Console.PrintError(Input + "\n")
        elif Type == "Log":
            App.Console.PrintLog(Input + "\n")
        else:
            App.Console.PrintMessage(Input + "\n")

    @classmethod
    def LightOrDark(rgbColor=[0, 128, 255, 255]):
        """_summary_
        reference: https://alienryderflex.com/hsp.html
        Args:
            rgbColor (list, optional): RGB color. Defaults to [0, 128, 255, 255].\n
            note: The alpha value is added for completeness, but us ignored in the equation.

        Returns:
            string: "light or dark"
        """
        import math

        [r, g, b, a] = rgbColor
        hsp = math.sqrt(0.299 * (r * r) + 0.587 * (g * g) + 0.114 * (b * b))
        if hsp > 127.5:
            return "light"
        else:
            return "dark"

    @classmethod
    def toggleToolbars(ToolbarName: str, WorkBench: str = ""):
        import FreeCADGui as Gui
        from PySide.QtWidgets import QToolBar

        # Get the active workbench
        if WorkBench == "":
            WB = Gui.activeWorkbench()
        if WorkBench != "":
            WB = Gui.getWorkbench(WorkBench)

        # Get the list of toolbars present.
        ListToolbars = WB.listToolbars()
        # Go through the list. If the toolbar exists set ToolbarExists to True
        ToolbarExists = False
        for i in range(len(ListToolbars)):
            if ListToolbars[i] == ToolbarName:
                ToolbarExists = True

        # If ToolbarExists is True continue. Otherwise return.
        if ToolbarExists is True:
            # Get the main window
            mainWindow = Gui.getMainWindow()
            # Get the toolbar
            ToolBar = mainWindow.findChild(QToolBar, ToolbarName)
            # If the toolbar is not hidden, hide it and return.
            if ToolBar.isHidden() is False:
                ToolBar.setHidden(True)
                return
            # If the toolbar is hidden, set visible and return.
            if ToolBar.isHidden() is True:
                ToolBar.setVisible(True)
                return
        return


    def PartFeatureList():
        result = [
            "Part::PartFeature",
            "Part::Extrusion",
            "Part::CustomFeature",
            "Part::Chamfer",
            "Part::Compound",
            "Part::Fillet",
            "Part::FeatureGeometrySet",
            "Part::Mirroring",
            "Part::Boolean",
            "Part::Box",
            "Part::Circle",
            "Part::MultiCommon",
            "Part::CurveNet",
            "Part::MultiFuse",
            "Part::ImportBrep",
            "Part::ImportIges",
            "Part::ImportStep",
            "Part::Polygon",
            "Part::RuledSurface",
            "Part::Loft",
            "Part::Sweep",
            "Part::Thickness",
            "Part::Refine",
            "Part::Reverse",
            "Part::Vertex",
            "Part::Line",
            "Part::Plane",
            "Part::Sphere",
            "Part::Ellipsoid",
            "Part::Cylinder",
            "Part::Prism",
            "Part::RegularPolygon",
            "Part::Cone",
            "Part::Torus",
            "Part::Helix",
            "Part::Spiral",
            "Part::Wedge",
            "Part::Ellipse",
            "Part::Extrusion",
            "Part::Face",
            "Part::Offset",
            "Part::Offset2D",
            "Part::Revolution",
            "Part::Part2DObject",
            "Part::Primitive",
        ]

        return result


    def PartDesingFeatureList():
        result = [
            "PartDesign::Pad",
            "PartDesign::Pocket",
            "PartDesign::FeaturePrimitive",
            "PartDesign::FeatureBase",
            "PartDesign::Draft",
            "PartDesign::Fillet",
            "PartDesign::Groove",
            "PartDesign::Helix",
            "PartDesign::Hole",
            "PartDesign::Loft",
            "PartDesign::Pipe",
            "PartDesign::Box",
            "PartDesign::Cylinder",
            "PartDesign::Sphere",
            "PartDesign::Cone",
            "PartDesign::Ellipsoid",
            "PartDesign::Torus",
            "PartDesign::Prism",
            "PartDesign::Wedge",
            "PartDesign::Revolution",
            "PartDesign::Thickness",
            "PartDesign::Transformed",
            "PartDesign::Body",
            "PartDesign::Boolean",
            "PartDesign::Chamfer",
            "PartDesign::ShapeBinder",
            "PartDesign::SubShapeBinder",
        ]

        return result

    # https://freecad.github.io/SourceDoc/d7/d7e/classPart_1_1Feature.html
    def AllFeaturesList():
        result = [
            "PartDesign::Pad",
            "PartDesign::Pocket",
            "PartDesign::FeaturePrimitive",
            "Drawing::FeatureProjection",
            "Part::CustomFeature",
            "Part::Chamfer",
            "Part::Compound",
            "Part::Fillet",
            "Part::FeatureGeometrySet",
            "Part::Mirroring",
            "Part::Boolean",
            "Part::Box",
            "Part::Circle",
            "Part::MultiCommon",
            "Part::CurveNet",
            "Part::MultiFuse",
            "Part::ImportBrep",
            "Part::ImportIges",
            "Part::ImportStep",
            "Part::Polygon",
            "Part::RuledSurface",
            "Part::Loft",
            "Part::Sweep",
            "Part::Thickness",
            "Part::Refine",
            "Part::Reverse",
            "Part::Vertex",
            "Part::Line",
            "Part::Plane",
            "Part::Sphere",
            "Part::Ellipsoid",
            "Part::Cylinder",
            "Part::Prism",
            "Part::RegularPolygon",
            "Part::Cone",
            "Part::Torus",
            "Part::Helix",
            "Part::Spiral",
            "Part::Wedge",
            "Part::Ellipse",
            "PartDesign::FeatureBase",
            "PartDesign::Draft",
            "PartDesign::Fillet",
            "PartDesign::Groove",
            "PartDesign::Helix",
            "PartDesign::Hole",
            "PartDesign::Loft",
            "PartDesign::Pipe",
            "PartDesign::Box",
            "PartDesign::Cylinder",
            "PartDesign::Sphere",
            "PartDesign::Cone",
            "PartDesign::Ellipsoid",
            "PartDesign::Torus",
            "PartDesign::Prism",
            "PartDesign::Wedge",
            "PartDesign::Revolution",
            "PartDesign::Thickness",
            "PartDesign::Transformed",
            "Path::FeatureArea",
            "Path::FeatureAreaView",
            "Sketcher::SketchObjectSF",
            "Surface::Cut",
            "Surface::Filling",
            "Surface::GeomFillSurface",
            "Surface::Sewing",
            "TechDraw::FeatureProjection",
            "Part::Extrusion",
            "Part::Face",
            "Part::Offset",
            "Part::Offset2D",
            "Part::Revolution",
            "Part::Part2DObject",
            "Part::Primitive",
            "PartDesign::Body",
            "PartDesign::Boolean",
            "PartDesign::Chamfer",
            "PartDesign::ShapeBinder",
            "PartDesign::SubShapeBinder",
            "Sketcher::SketchObject",
            "Surface::Extend",
            "Surface::Sections.",
        ]

        return result
